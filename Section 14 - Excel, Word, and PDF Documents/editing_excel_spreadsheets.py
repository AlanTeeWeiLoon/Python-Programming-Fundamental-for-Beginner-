import openpyxl

print('\n1---------------------------------\n')
# to create a new blank excel worksheet

wb = openpyxl.Workbook()
print(wb) # <openpyxl.workbook.workbook.Workbook object at 0x03446250>

print(wb.get_sheet_names()) # ['Sheet']

sheet = wb.get_sheet_by_name('Sheet')
print(sheet) # <Worksheet "Sheet">

a = sheet['A1'].value == None
print(a) # True - there is no value inside the worksheet

print('\n2---------------------------------\n')
# to save this worksheet as a excel file

sheet['A1'] = 42
sheet['A2'] = 'Hello'
wb.save('example1.xlsx') # the file saved with name 'example1.xlsx'

print('\n3---------------------------------\n')
# .create_sheet() - create a new sheet

sheet2 = wb.create_sheet()
print(wb.get_sheet_names()) # ['Sheet', 'Sheet1'] - a new sheet created

print('\n4---------------------------------\n')
# .title - change sheet name

print(sheet2.title) # Sheet1 - the name of sheet that created just now is 'Sheet1'

sheet2.title = 'My New Sheet Name'
print(wb.get_sheet_names()) # ['Sheet', 'My New Sheet Name'] - the sheet name have changed
wb.save('example2.xlsx') # save again to change the sheet name in excel file

print('\n5---------------------------------\n')
# to save a sheet to an specify place and with a name

wb.create_sheet(index = 0, title = 'My Other Sheet')
wb.save('example3.xlsx') # excel file saved with name 'example3.xlsx'












