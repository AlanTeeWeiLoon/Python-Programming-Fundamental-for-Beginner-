import openpyxl
import os

print('\n1--------------------------------------------------\n')
# openpyxl.load_workbook() - open excel file

workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook)) # <class 'openpyxl.workbook.workbook.Workbook'>

print('\n2--------------------------------------------------\n')
# workbook.get_sheet_by_name()

sheet = workbook.get_sheet_by_name('Sheet1') # if you know the name of sheet
print(type(sheet)) # <class 'openpyxl.worksheet.worksheet.Worksheet'>

print('\n3--------------------------------------------------\n')
# workbook.get_sheet_names() - Get sheets' name

getsheetname = workbook.get_sheet_names()
print(getsheetname) # ['Sheet1', 'Sheet2', 'Sheet3']

print('\n4--------------------------------------------------\n')
# Get the cell value

cell = sheet['A1']
print(cell.value) # 2015-04-05 13:34:02

cell1 = sheet['B1']
print(cell1.value) # Apples

cell2 = sheet['C1']
print(cell2.value) # 73

print('\n5--------------------------------------------------\n')
# test row and column - return the same value with using cell name

a = sheet.cell(row = 1, column = 2)
print(a) # <Cell 'Sheet1'.B1>

b = sheet['B1']
print(b) # <Cell 'Sheet1'.B1>

print('\n6--------------------------------------------------\n')
# using loop with the value in the excel file

for i in range(1, 8):
    print(i, sheet.cell(row = 1, column = 2).value) 

# 1 Apples
# 2 Apples
# 3 Apples
# 4 Apples
# 5 Apples
# 6 Apples
# 7 Apples

























